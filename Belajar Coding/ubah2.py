import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from io import BytesIO

def clean_data(df):
    df.columns = df.columns.str.strip()
    kolom_disimpan = ['Nama produk', 'Nama varian', 'SKU', 'Total Kuantitas']
    df_filtered = df[kolom_disimpan]
    df_unique = df_filtered.drop_duplicates()
    df_unique.loc[:, 'Total Kuantitas'] = pd.to_numeric(df_unique['Total Kuantitas'], errors='coerce').fillna(0).astype(int)
    df_unique = df_unique.rename(columns={'Total Kuantitas': 'Qty'})
    df_sorted = df_unique.sort_values(by='Qty', ascending=False)
    return df_sorted

def format_and_save_to_excel(df):
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment

    # Lebar kolom
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 4
    
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output


def main(file_obj):
    """
    Menerima file upload (Flask FileStorage), proses, dan return hasil sebagai BytesIO
    """
    df = pd.read_excel(file_obj)
    cleaned_df = clean_data(df)
    final_file = format_and_save_to_excel(cleaned_df)
    return final_file
