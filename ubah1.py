import pdfplumber
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

def extract_and_process_pdf(pdf_path):
    print("Memulai metode 'Grid' untuk ekstraksi data mentah...")

    KOLOM_BOUNDARIES = [
        (0, 350), (350, 470), (470, 540), (540, 595)
    ]

    all_rows_structured = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            print(f"Memproses Halaman {page_num + 1}...")
            
            if page.width != 595:
                KOLOM_BOUNDARIES[3] = (540, page.width)

            h_lines = sorted(list(set([edge['top'] for edge in page.horizontal_edges] + [0, page.height])))

            for i in range(len(h_lines) - 1):
                top = h_lines[i]
                bottom = h_lines[i+1]
                
                row_data = []
                for x0, x1 in KOLOM_BOUNDARIES:
                    cell_crop = page.crop((x0, top, x1, bottom))
                    text = cell_crop.extract_text(x_tolerance=2, y_tolerance=2)
                    row_data.append(text.strip() if text else '')
                
                if any(row_data):
                    all_rows_structured.append(row_data)

    if not all_rows_structured:
        raise Exception("Tidak ada data yang bisa diekstrak.")

    df_raw = pd.DataFrame(all_rows_structured, columns=['Nama Produk', 'SKU', 'Slot', 'Qty'])
    df_raw = df_raw[~df_raw['Nama Produk'].str.contains("Nama Produk", na=False)].reset_index(drop=True)
    
    return df_raw


def clean_data(df_raw):
    """
    Fungsi yang diperbarui dengan logika Qty yang lebih kuat.
    """
    print("Menerapkan aturan pembersihan cerdas pada data...")
    processed_data = []
    
    for i in range(len(df_raw)):
        current_row = df_raw.iloc[i]
        qty_raw = str(current_row.get('Qty', ''))

        # --- PERBAIKAN UTAMA DI SINI ---
        # Cari rangkaian angka pertama di dalam string Qty yang mungkin kotor
        qty_match = re.search(r'\d+', qty_raw)

        # Gunakan baris yang punya QTY sebagai "Jangkar" (Anchor)
        if qty_match:
            qty = qty_match.group(0) # Ambil angka yang bersih (misal: '32' dari '32\nman')

            # Inisialisasi semua bagian data
            nama_produk_final = ""
            varian_final = ""
            sku_final = str(current_row.get('SKU', ''))

            if current_row.get('Nama Produk'):
                nama_produk_final = str(current_row.get('Nama Produk'))
            else:
                if i > 0:
                    prev_row_1 = df_raw.iloc[i-1]
                    prev_text_1 = str(prev_row_1.get('Nama Produk', ''))
                    
                    if 'Variant:' in prev_text_1 or 'riant:' in prev_text_1:
                        varian_final = re.sub(r'(variant:|riant:)', '', prev_text_1, flags=re.IGNORECASE).strip()
                        if i > 1:
                            prev_row_2 = df_raw.iloc[i-2]
                            nama_produk_final = str(prev_row_2.get('Nama Produk', ''))
                    else:
                        nama_produk_final = prev_text_1
            
            if 'Buyer Notes:' in nama_produk_final:
                nama_produk_final = nama_produk_final.split('Buyer Notes:')[0]

            sku_joined = sku_final.replace('\n', '')
            sku_cleaned = re.sub('defa', '', sku_joined, flags=re.IGNORECASE).strip()
            sku_final_cleaned = re.sub(r'^.\s', '', sku_cleaned)
            
            nama_produk_clean = ' '.join(nama_produk_final.replace('\n', ' ').split())
            varian = ''
            match = re.search(r'(variant:|riant:)(.*)', nama_produk_clean, re.IGNORECASE)
            if match:
                nama_produk_clean = nama_produk_clean.split(match.group(0))[0].strip()
                varian = match.group(2).strip()
            
            nama_produk_terbatas = nama_produk_clean[:90]
                        
            processed_data.append({
                'Nama Produk': nama_produk_terbatas,
                'Varian': varian,
                'SKU': sku_final_cleaned,
                'Qty': int(qty)
            })
            
    return processed_data


def save_to_excel_in_memory(data):
    wb = Workbook()
    ws = wb.active

    headers = ["Nama Produk", "Varian", "SKU", "Qty"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for item in data:
        row = [
            item.get("Nama Produk", ""),
            item.get("Varian", ""),
            item.get("SKU", ""),
            item.get("Qty", ""),
        ]
        ws.append(row)

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 4

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
            
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def main(file_path):
    
    raw_data_df = extract_and_process_pdf(file_path)
    cleaned_data = clean_data(raw_data_df)
    excel_file_in_memory = save_to_excel_in_memory(cleaned_data)

    print("\nProses Selesai! File Excel dibuat di memori.")
    return excel_file_in_memory
